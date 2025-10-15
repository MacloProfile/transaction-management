import csv
from io import TextIOWrapper
from openpyxl import load_workbook


def iter_csv(uploaded_file, encoding='utf-8-sig'):
    wrapper = TextIOWrapper(uploaded_file.file, encoding=encoding, newline='')
    reader = csv.reader(wrapper, delimiter=';')

    headers = None
    for i, row in enumerate(reader):
        if not any(cell.strip() for cell in row):
            continue
        if i == 0:
            headers = [h.strip().lower() for h in row]
            continue
        yield row_to_dict(row)


def row_to_dict(row):
    return {
        'first_name': row[0].strip() if len(row) > 0 else '',
        'last_name': row[1].strip() if len(row) > 1 else '',
        'phone': str(row[2]).strip() if len(row) > 2 else '',
        'email': row[3].strip() if len(row) > 3 else '',
        'company': row[4].strip() if len(row) > 4 else '',
    }


def iter_xlsx(uploaded_file):
    wb = load_workbook(filename=uploaded_file, read_only=True)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not any(row):
            continue
        if i == 0:
            headers = [(cell or '').strip().lower() for cell in row]
            continue
        yield row_to_dict([(cell or '') for cell in row])
